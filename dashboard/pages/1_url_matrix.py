"""情報源URLマトリクス — url_matrix.db ベース版

機関 × カテゴリ × 入札方式 × 月 のセルで status と URL を可視化。
- filled: 既収集確認済（緑）
- filled_new: 新規充填（青）
- dash: URL 未設定（赤）
- flag_collected=1: procurement.db.source_url との照合一致
- flag_404=1: 取得失敗
"""
from __future__ import annotations

import io
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import streamlit as st

from _auth import require_password
require_password()

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
URL_DB = PROJECT_ROOT / "data" / "db" / "url_matrix.db"
PROC_DB = PROJECT_ROOT / "data" / "db" / "procurement.db"

CAT_LARGE_ORDER: dict[str, int] = {
    "内部部局": 0, "統合幕僚監部": 1, "情報本部": 2,
    "陸上自衛隊": 3, "海上自衛隊": 4, "航空自衛隊": 5,
    "地方防衛局": 6, "防衛大学校": 7, "防衛研究所": 8,
    "防衛医科大学校": 9, "防衛監察本部": 10, "防衛装備庁": 11,
}

TEXT_COLOR = "#cdd6f4"
TEXT_DIM = "#bac2de"
ACCENT = "#7c83fd"

st.markdown(f"""
<style>
.stApp, .stApp p, .stApp span, .stApp label {{ color: {TEXT_COLOR}; }}
.stApp h1, .stApp h2, .stApp h3 {{ color: {TEXT_COLOR}; }}
[data-testid="stCaptionContainer"] {{ color: {TEXT_DIM} !important; }}
[data-testid="stMetricValue"] {{ color: {TEXT_COLOR}; font-weight: 600; }}
[data-testid="stMetricLabel"] {{ color: {TEXT_DIM}; font-size: 0.85rem; }}
a, a:visited {{ color: {ACCENT}; text-decoration: none; }}
a:hover {{ color: #a5b4fc; text-decoration: underline; }}

table.site-url-table {{
    border-collapse: collapse;
    font-size: 0.78rem;
    width: 100%;
}}
table.site-url-table th, table.site-url-table td {{
    border: 1px solid #313244;
    padding: 4px 8px;
    vertical-align: middle;
    color: {TEXT_COLOR};
}}
table.site-url-table th {{
    background: #1e1e2e;
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 2;
    white-space: nowrap;
}}
table.site-url-table td.no-cell {{
    text-align: right;
    color: {TEXT_DIM};
    width: 36px;
    font-size: 0.72rem;
}}
table.site-url-table td.name-cell {{
    background: #181825;
    font-weight: 600;
    min-width: 140px;
    max-width: 180px;
}}
table.site-url-table td.url-cell {{
    min-width: 160px;
    max-width: 260px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}}
table.site-url-table td.warp-cell {{
    text-align: center;
    min-width: 70px;
    white-space: nowrap;
}}
table.site-url-table td.warp-hit  {{ background: #1a2a1a; }}
table.site-url-table td.warp-none {{ background: #2a1a1a; color: #585b70; }}
table.site-url-table a.url-link {{
    color: {ACCENT};
    text-decoration: none;
    display: block;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    max-width: 250px;
    font-size: 0.75rem;
}}
table.site-url-table a.warp-link {{
    color: #a6e3a1;
    font-size: 0.72rem;
    text-decoration: none;
}}
table.site-url-table a.url-link:hover,
table.site-url-table a.warp-link:hover {{ text-decoration: underline; }}

table.url-matrix {{
    border-collapse: collapse;
    font-size: 0.75rem;
    width: 100%;
}}
table.url-matrix th, table.url-matrix td {{
    border: 1px solid #313244;
    padding: 3px 6px;
    white-space: nowrap;
    vertical-align: middle;
    text-align: center;
    color: {TEXT_COLOR};
}}
table.url-matrix th {{
    background: #1e1e2e;
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 2;
}}
table.url-matrix td.agency-cell {{
    background: #181825;
    font-weight: 600;
    text-align: left;
    min-width: 110px;
    max-width: 150px;
    overflow: hidden;
    text-overflow: ellipsis;
}}
table.url-matrix td.cat-cell {{
    background: #1e1e2e;
    text-align: left;
    width: 80px;
    font-size: 0.72rem;
}}
/* 状態別カラー */
table.url-matrix td.s-filled       {{ background: #1a2a1a; }}      /* 緑：既存filled+collected */
table.url-matrix td.s-filled-new   {{ background: #1a253a; }}      /* 青：今回新規充填+collected */
table.url-matrix td.s-pending      {{ background: #3a3210; }}      /* 黄：URLあり未確認 */
table.url-matrix td.s-404          {{ background: #3a1818; color: #f38ba8; }}  /* 暗赤：404 */
table.url-matrix td.s-dash         {{ background: #2a1a1a; color: #585b70; }}  /* 赤：URL未設定 */
table.url-matrix td.s-dash::after  {{ content: "－"; color: #45475a; }}
table.url-matrix a {{
    display: block;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 90px;
    font-size: 0.72rem;
    color: {ACCENT};
}}
.legend-chip {{
    display: inline-block; padding: 2px 8px; margin-right: 6px;
    border-radius: 3px; font-size: 0.75rem; color: {TEXT_COLOR};
    border: 1px solid #313244;
}}
/* 大区分/中区分グループヘッダー行 */
table.url-matrix tr.group-large td,
table.site-url-table tr.group-large td {{
    background: #11111b;
    font-weight: 700;
    font-size: 0.82rem;
    color: #cba6f7;
    padding: 5px 8px;
    border-top: 2px solid #45475a;
    text-align: left;
}}
table.url-matrix tr.group-mid td,
table.site-url-table tr.group-mid td {{
    background: #181825;
    font-weight: 600;
    font-size: 0.76rem;
    color: #89b4fa;
    padding: 3px 8px 3px 18px;
    text-align: left;
}}
</style>
""", unsafe_allow_html=True)

st.title("📊 情報源URLマトリクス")
st.caption("機関 × カテゴリ × 入札方式 × 月で表示。")

if not URL_DB.exists():
    st.error(f"DBが見つかりません: {URL_DB}")
    st.stop()


# ── データロード ────────────────────────────────────────
@st.cache_data(ttl=120)
def _load_url_matrix(fy: int = 2024) -> pd.DataFrame:
    with sqlite3.connect(URL_DB) as conn:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(url_matrix)")}
        sort_col = "sort_order," if "sort_order" in cols else ""
        if "fiscal_year" in cols:
            where = f"WHERE fiscal_year = {int(fy)}"
        else:
            where = ""
        df = pd.read_sql_query(
            f"""
            SELECT id, agency_id, agency_name, category, bid_method, month,
                   filename, url, status,
                   {sort_col}
                   IFNULL(flag_rolling, 0)     AS flag_rolling,
                   IFNULL(flag_warp, 0)        AS flag_warp,
                   IFNULL(flag_mixed_cat, 0)   AS flag_mixed_cat,
                   IFNULL(flag_mixed_month, 0) AS flag_mixed_month,
                   IFNULL(flag_collected, 0)   AS flag_collected,
                   IFNULL(flag_404, 0)         AS flag_404
            FROM url_matrix {where}
            """,
            conn,
        )
    if "sort_order" not in df.columns:
        df["sort_order"] = None
    return df


@st.cache_data(ttl=120)
def _load_site_urls() -> pd.DataFrame:
    """agency_id 単位の掲載サイトURL + WARPスナップショットURLを返す。"""
    with sqlite3.connect(URL_DB) as conn:
        # site_url カラムが存在するか確認
        cols = {r[1] for r in conn.execute("PRAGMA table_info(url_matrix)")}
        if "site_url" not in cols:
            return pd.DataFrame()
        warp_cols = [c for c in ["site_url_2025_07", "site_url_2024_07", "site_url_2023_07"] if c in cols]
        warp_sel = ", ".join(f"MAX({c}) AS {c}" for c in warp_cols) if warp_cols else ""
        sort_sel = ", MIN(sort_order) AS sort_order" if "sort_order" in cols else ""
        sel = f"MAX(agency_name) AS agency_name, MAX(site_url) AS site_url"
        if warp_sel:
            sel += f", {warp_sel}"
        df = pd.read_sql_query(
            f"SELECT agency_id, {sel}{sort_sel} FROM url_matrix GROUP BY agency_id ORDER BY MAX(agency_name)",
            conn,
        )
    if "sort_order" not in df.columns:
        df["sort_order"] = None
    return df


# agency_id の表記ゆれ・追加機関に対するカテゴリ手動オーバーライド
# procurement.db に登録がない / 別IDで登録されている機関を補完
CATEGORY_OVERRIDES: dict[str, tuple[str, str]] = {
    "gsdf_eadep_yoga":  ("陸上自衛隊", ""),  # procurement.db は yooga（oo）表記
    "gsdf_eadep_koga":  ("陸上自衛隊", ""),
    "gsdf_eadep":       ("陸上自衛隊", ""),
    "gsdf_eafin":       ("陸上自衛隊", ""),
    "nda":              ("防衛大学校", ""),
    "igo":              ("防衛監察本部", ""),
}



@st.cache_data(ttl=300)
def _load_category_map() -> pd.DataFrame:
    """procurement.db から agency_id → (category_large, category_mid) を取得。"""
    if not PROC_DB.exists():
        return pd.DataFrame(columns=["agency_id", "category_large", "category_mid"])
    with sqlite3.connect(PROC_DB) as conn:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(contracts)")}
        if "category_large" not in cols:
            return pd.DataFrame(columns=["agency_id", "category_large", "category_mid"])
        df = pd.read_sql_query(
            """SELECT agency_id,
                      MAX(category_large) AS category_large,
                      MAX(COALESCE(NULLIF(category_mid,''), NULL)) AS category_mid
               FROM contracts
               WHERE category_large IS NOT NULL AND category_large != ''
               GROUP BY agency_id""",
            conn,
        )
    # オーバーライド行を追加（procurement.db に存在しない / NULLのもの）
    existing = set(df["agency_id"])
    extras = [
        {"agency_id": aid, "category_large": lg, "category_mid": mid}
        for aid, (lg, mid) in CATEGORY_OVERRIDES.items()
        if aid not in existing
    ]
    if extras:
        df = pd.concat([df, pd.DataFrame(extras)], ignore_index=True)
    # 既存行でも category_large が空なら上書き
    for aid, (lg, mid) in CATEGORY_OVERRIDES.items():
        mask = (df["agency_id"] == aid) & (df["category_large"].isna() | (df["category_large"] == ""))
        df.loc[mask, "category_large"] = lg
        df.loc[mask, "category_mid"]   = mid
    return df


def _branch(aid: str) -> str:
    if not aid:
        return "その他"
    for prefix, label in [
        ("atla", "ATLA"), ("asdf", "ASDF"), ("msdf", "MSDF"),
        ("gsdf", "GSDF"), ("rdb", "RDB"), ("ndmc", "NDMC"),
        ("nida", "防衛大"), ("js", "統幕"), ("dih", "情本"),
        ("naikyoku", "内局"),
    ]:
        if aid.startswith(prefix):
            return label
    return "その他"


def _cell_state(row: pd.Series) -> str:
    """セル分類: 'filled' / 'filled_new' / 'pending' / '404' / 'dash'"""
    status = (row.get("status") or "").strip()
    has_url = bool(row.get("url"))
    if not has_url:
        return "dash"
    if int(row.get("flag_404") or 0) == 1:
        return "404"
    collected = int(row.get("flag_collected") or 0) == 1
    if status == "filled" and collected:
        return "filled"
    if status == "filled_new" and collected:
        return "filled_new"
    return "pending"


selected_fy = st.selectbox(
    "年度 (FY)",
    options=[2025, 2024, 2023, 2022],
    index=1,  # default: FY2024 (最高充填率)
    key="um_fy",
    help="選択したFYのセル情報を表示。各FYとも収集済 source_url を url 列に反映済み。",
)

with st.spinner("url_matrix.db 読み込み中..."):
    df = _load_url_matrix(selected_fy)
    cat_map = _load_category_map()

if df.empty:
    st.warning("url_matrix テーブルが空です。")
    st.stop()

df = df.merge(cat_map, on="agency_id", how="left")
df["category_large"] = df["category_large"].fillna("その他")
df["category_mid"]   = df["category_mid"].fillna("")
df["branch"] = df["agency_id"].fillna("").map(_branch)
df["state"]  = df.apply(_cell_state, axis=1)


# ── KPI ─────────────────────────────────────────────────
total = len(df)
with_url = int((df["url"].notna() & (df["url"] != "")).sum())
collected = int((df["flag_collected"] == 1).sum())
n_filled = int((df["state"] == "filled").sum())
n_filled_new = int((df["state"] == "filled_new").sum())
n_pending = int((df["state"] == "pending").sum())
n_404 = int((df["state"] == "404").sum())
n_dash = int((df["state"] == "dash").sum())

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("総セル数", f"{total:,}")
k2.metric("URL設定済", f"{with_url:,}", delta=f"{with_url/total*100:.1f}%")
k3.metric("既収集確認", f"{collected:,}", delta=f"{collected/total*100:.1f}%")
k4.metric("新規充填", f"{n_filled_new:,}")
k5.metric("URL未設定 (dash)", f"{n_dash:,}", delta=f"{n_dash/total*100:.1f}%",
          delta_color="inverse")

st.markdown(
    '<div style="margin: 4px 0 12px 0;">'
    f'<span class="legend-chip" style="background:#1a2a1a">緑: 既存収集済 ({n_filled})</span>'
    f'<span class="legend-chip" style="background:#1a253a">青: 新規充填 ({n_filled_new})</span>'
    f'<span class="legend-chip" style="background:#3a3210">黄: URLあり未確認 ({n_pending})</span>'
    f'<span class="legend-chip" style="background:#3a1818;color:#f38ba8">暗赤: 404 ({n_404})</span>'
    f'<span class="legend-chip" style="background:#2a1a1a;color:#585b70">赤: URL未設定 ({n_dash})</span>'
    '</div>',
    unsafe_allow_html=True,
)

st.divider()


# ── サマリ：カテゴリ × 入札方式別の充填率 ─────────────────
st.subheader("📈 充填率サマリ（カテゴリ × 入札方式）")

summary = (
    df.assign(
        is_filled=df["state"].isin(["filled", "filled_new"]).astype(int),
        is_url=(df["url"].notna() & (df["url"] != "")).astype(int),
    )
    .groupby(["category", "bid_method"], as_index=False)
    .agg(
        セル数=("id", "count"),
        URL設定済=("is_url", "sum"),
        収集済=("is_filled", "sum"),
    )
)
summary["URL率"] = (summary["URL設定済"] / summary["セル数"] * 100).round(1)
summary["収集率"] = (summary["収集済"] / summary["セル数"] * 100).round(1)
summary = summary.rename(columns={"category": "カテゴリ", "bid_method": "入札方式"})
summary["URL率"] = summary["URL率"].map(lambda v: f"{v:.1f}%")
summary["収集率"] = summary["収集率"].map(lambda v: f"{v:.1f}%")

# ブランチ別サマリ
branch_summary = (
    df.assign(
        is_filled=df["state"].isin(["filled", "filled_new"]).astype(int),
        is_url=(df["url"].notna() & (df["url"] != "")).astype(int),
    )
    .groupby("branch", as_index=False)
    .agg(
        セル数=("id", "count"),
        URL設定済=("is_url", "sum"),
        収集済=("is_filled", "sum"),
    )
)
branch_summary["URL率"] = (branch_summary["URL設定済"] / branch_summary["セル数"] * 100).round(1)
branch_summary["収集率"] = (branch_summary["収集済"] / branch_summary["セル数"] * 100).round(1)
branch_summary = branch_summary.sort_values("収集済", ascending=False)
branch_summary["URL率"] = branch_summary["URL率"].map(lambda v: f"{v:.1f}%")
branch_summary["収集率"] = branch_summary["収集率"].map(lambda v: f"{v:.1f}%")
branch_summary = branch_summary.rename(columns={"branch": "ブランチ"})

sum_c1, sum_c2 = st.columns(2)
with sum_c1:
    st.markdown("**カテゴリ × 入札方式**")
    st.dataframe(summary, use_container_width=True, hide_index=True, height=200)
with sum_c2:
    st.markdown("**ブランチ別**")
    st.dataframe(branch_summary, use_container_width=True, hide_index=True, height=320)

st.divider()


# ── フィルタ ─────────────────────────────────────────────
st.subheader("🔍 フィルタ")

all_large = sorted(df["category_large"].unique(), key=lambda x: CAT_LARGE_ORDER.get(x, 99))
all_mid   = sorted(df["category_mid"].dropna().unique())
all_cats  = sorted(df["category"].dropna().unique())
all_bids  = sorted(df["bid_method"].dropna().unique())
all_states = ["filled", "filled_new", "pending", "404", "dash"]

fr1, fr2 = st.columns([3, 3])
with fr1:
    sel_large = st.multiselect("大区分", all_large, default=all_large, key="um_large")
with fr2:
    sel_mid = st.multiselect("中区分", all_mid, default=all_mid, key="um_mid",
                             help="空欄は中区分なし（本部・単独機関）")

fr3, fr4, fr5 = st.columns([2, 2, 2])
with fr3:
    sel_cat = st.multiselect("カテゴリ", all_cats, default=all_cats, key="um_cat")
with fr4:
    sel_bid = st.multiselect("入札方式", all_bids, default=all_bids, key="um_bid")
with fr5:
    sel_state = st.multiselect(
        "状態", all_states, default=all_states, key="um_state",
        help="filled=既存緑 / filled_new=新規青 / pending=URLあり未確認 / 404=取得失敗 / dash=URL未設定",
    )

mdf = df[
    df["category_large"].isin(sel_large)
    & (df["category_mid"].isin(sel_mid) | (df["category_mid"] == ""))
    & df["category"].isin(sel_cat)
    & df["bid_method"].isin(sel_bid)
    & df["state"].isin(sel_state)
].copy()

# state フィルタを除いたビュー（row_keys生成用: 全4コンボを常に表示するため）
mdf_nostate = df[
    df["category_large"].isin(sel_large)
    & (df["category_mid"].isin(sel_mid) | (df["category_mid"] == ""))
    & df["category"].isin(sel_cat)
    & df["bid_method"].isin(sel_bid)
].copy()

if mdf_nostate.empty:
    st.info("条件に一致するセルがありません。フィルタを変更してください。")
    st.stop()

st.caption(f"絞り込み後: {len(mdf):,}セル / {mdf_nostate['agency_id'].nunique()} 機関")

# ── マトリクス本体 ───────────────────────────────────────
st.divider()
st.subheader(f"🧮 マトリクス（FY{selected_fy}）")

FY_MONTHS = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
N_MATRIX_COLS = 3 + len(FY_MONTHS)  # No. + 機関 + カテゴリ/入札方式 + 12ヶ月

# 4標準コンボ（表示用固定）
STANDARD_COMBOS = [
    ("公共工事", "競争"),
    ("公共工事", "随契"),
    ("物品役務", "競争"),
    ("物品役務", "随契"),
]

# cell_index: 表示対象機関の全行（state非依存でセル色付けに使用）
_visible_aids = set(mdf_nostate["agency_id"])
cell_index: dict[tuple, pd.Series] = {}
for _, r in df[df["agency_id"].isin(_visible_aids)].iterrows():
    cell_index[
        (r["agency_id"], r["agency_name"], r["category"], r["bid_method"], int(r["month"]))
    ] = r

# 機関メタ情報（表示対象機関ごと1行）
agency_meta = (
    mdf_nostate.groupby("agency_id")
    .agg(
        agency_name=("agency_name", "first"),
        category_large=("category_large", "first"),
        category_mid=("category_mid", "first"),
        sort_order=("sort_order", "min"),
    )
    .reset_index()
)
_sort_order_map = dict(zip(agency_meta["agency_id"], agency_meta["sort_order"]))

# 全4コンボ × 機関 の row_keys（DBに行がなくても必ず4行表示）
row_keys_raw = []
for _, ag in agency_meta.iterrows():
    aid       = ag["agency_id"]
    aname     = ag["agency_name"]
    cat_large = ag["category_large"]
    cat_mid   = ag["category_mid"]
    for cat, bid in STANDARD_COMBOS:
        if cat in sel_cat and bid in sel_bid:
            row_keys_raw.append((aid, aname, cat, bid, cat_large, cat_mid))

row_keys = sorted(
    row_keys_raw,
    key=lambda x: (
        CAT_LARGE_ORDER.get(x[4], 99), x[4],
        _sort_order_map.get(x[0], 9999),
        x[5], x[1], x[2], x[3],
    ),
)

st.caption(f"{len(row_keys)} 行 × 12 ヶ月")

header_cells = (
    '<th style="text-align:right;width:36px">No.</th>'
    '<th style="text-align:left;min-width:140px">機関</th>'
    '<th style="text-align:left;min-width:80px">カテゴリ<br>入札方式</th>'
    + "".join(f'<th>{m}月</th>' for m in FY_MONTHS)
)

html_rows: list[str] = []
prev_aid = prev_large = prev_mid = None
row_no = 0
for aid, aname, cat, bid, cat_large, cat_mid in row_keys:
    # 大区分ヘッダー行
    if cat_large != prev_large:
        html_rows.append(
            f'<tr class="group-large">'
            f'<td colspan="{N_MATRIX_COLS}">▌ {cat_large}</td></tr>'
        )
        prev_large = cat_large
        prev_mid = None  # 大区分が変わったら中区分もリセット
    # 中区分ヘッダー行
    if cat_mid and cat_mid != prev_mid:
        html_rows.append(
            f'<tr class="group-mid">'
            f'<td colspan="{N_MATRIX_COLS}">　└ {cat_mid}</td></tr>'
        )
        prev_mid = cat_mid

    row_no += 1
    no_td = (
        f'<td style="text-align:right;color:{TEXT_DIM};'
        f'font-size:0.72rem;padding:3px 4px">{row_no}</td>'
    )
    if aid != prev_aid:
        agency_td = f'<td class="agency-cell" title="{aid}">{aname}</td>'
        prev_aid = aid
    else:
        agency_td = '<td class="agency-cell" style="border-top:none"></td>'
    cat_td = (
        f'<td class="cat-cell">{cat}<br>'
        f'<span style="color:{TEXT_DIM}">{bid}</span></td>'
    )

    month_tds: list[str] = []
    for mm in FY_MONTHS:
        rec = cell_index.get((aid, aname, cat, bid, mm))
        if rec is None:
            month_tds.append('<td class="s-dash"></td>')
            continue
        state = rec["state"]
        if state == "dash" or state not in sel_state:
            month_tds.append('<td class="s-dash"></td>')
        else:
            cls = {
                "filled": "s-filled",
                "filled_new": "s-filled-new",
                "pending": "s-pending",
                "404": "s-404",
            }[state]
            url = rec["url"] or ""
            fname = (rec["filename"] or url.rsplit("/", 1)[-1] if url else "?")[:35]
            tip = f"status={rec['status']} state={state}"
            if rec.get("flag_warp"):
                tip += " [WARP]"
            if rec.get("flag_rolling"):
                tip += " [Rolling]"
            if url:
                cell = (
                    f'<td class="{cls}" title="{tip}">'
                    f'<a href="{url}" target="_blank">{fname}</a></td>'
                )
            else:
                cell = f'<td class="{cls}" title="{tip}">{fname}</td>'
            month_tds.append(cell)

    html_rows.append(
        "<tr>" + no_td + agency_td + cat_td + "".join(month_tds) + "</tr>"
    )

st.markdown(
    '<div style="overflow-x:auto;max-height:72vh;overflow-y:auto;">'
    f'<table class="url-matrix"><thead><tr>{header_cells}</tr></thead>'
    f'<tbody>{"".join(html_rows)}</tbody></table></div>',
    unsafe_allow_html=True,
)


# ── 掲載サイトURL & WARPスナップショット ────────────────────
st.divider()
st.subheader("🌐 掲載サイトURL & WARPスナップショット一覧")
st.caption("各機関の掲載サイトURL（現行）と過去3時点のWARPアーカイブURL。")

site_df = _load_site_urls()
if site_df.empty:
    st.info("site_url カラムが未設定です。setup_categories.py を実行してください。")
else:
    site_df = site_df.merge(cat_map, on="agency_id", how="left")
    site_df["category_large"] = site_df["category_large"].fillna("その他")
    site_df["category_mid"]   = site_df["category_mid"].fillna("")

    # 大区分フィルタ（マトリクス側と共有）
    sb_large = sorted(site_df["category_large"].unique(),
                      key=lambda x: CAT_LARGE_ORDER.get(x, 99))
    sel_sb = st.multiselect(
        "大区分絞り込み", sb_large, default=sb_large, key="sb_large"
    )
    sdf = site_df[site_df["category_large"].isin(sel_sb)].copy()
    sdf["_sort_order"] = sdf["sort_order"].where(sdf["sort_order"].notna(), 9999)
    sdf = sdf.sort_values(
        by=["category_large", "_sort_order", "category_mid", "agency_name"],
        key=lambda col: col.map(lambda v: CAT_LARGE_ORDER.get(v, 99)) if col.name == "category_large" else col,
    ).drop(columns=["_sort_order"])
    st.caption(f"{len(sdf)} 機関")

    def _short(url: str | None, max_len: int = 55) -> str:
        if not url:
            return ""
        import re as _re
        m = _re.match(r"https://warp\.ndl\.go\.jp/\d+/[^/]+/(.+)", url)
        display = m.group(1) if m else url
        display = display.replace("https://www.mod.go.jp/", "").replace("https://", "")
        return display[:max_len] + ("…" if len(display) > max_len else "")

    def _warp_cell(url: str | None) -> str:
        if url:
            short = _short(url, 38)
            return (
                f'<td class="warp-cell warp-hit">'
                f'<a class="warp-link" href="{url}" target="_blank" title="{url}">✅ {short}</a></td>'
            )
        return '<td class="warp-cell warp-none">—</td>'

    warp_labels = [
        ("site_url_2025_07", "2025.7"),
        ("site_url_2024_07", "2024.7"),
        ("site_url_2023_07", "2023.7"),
    ]
    warp_header = "".join(f"<th>{lbl}</th>" for col, lbl in warp_labels if col in site_df.columns)
    N_SITE_COLS = 3 + sum(1 for col, _ in warp_labels if col in site_df.columns)
    hdr = (
        '<th class="no-cell">No.</th>'
        '<th>機関名</th>'
        '<th>掲載サイトURL（現行）</th>'
        + warp_header
    )

    rows_html: list[str] = []
    s_prev_large = s_prev_mid = None
    s_row_no = 0
    for _, row in sdf.iterrows():
        clarge = row.get("category_large") or "その他"
        cmid   = row.get("category_mid") or ""
        if clarge != s_prev_large:
            rows_html.append(
                f'<tr class="group-large"><td colspan="{N_SITE_COLS}">▌ {clarge}</td></tr>'
            )
            s_prev_large = clarge
            s_prev_mid = None
        if cmid and cmid != s_prev_mid:
            rows_html.append(
                f'<tr class="group-mid"><td colspan="{N_SITE_COLS}">　└ {cmid}</td></tr>'
            )
            s_prev_mid = cmid

        s_row_no += 1
        site_url = row.get("site_url") or ""
        site_cell = (
            f'<td class="url-cell"><a class="url-link" href="{site_url}" target="_blank"'
            f' title="{site_url}">{_short(site_url)}</a></td>'
            if site_url
            else '<td class="url-cell" style="color:#585b70">—</td>'
        )
        warp_cells = "".join(
            _warp_cell(row.get(col))
            for col, _ in warp_labels
            if col in site_df.columns
        )
        aid_val  = row["agency_id"]
        name_val = row["agency_name"] or row["agency_id"]
        rows_html.append(
            f"<tr>"
            f'<td class="no-cell">{s_row_no}</td>'
            f'<td class="name-cell" title="{aid_val}">{name_val}</td>'
            f"{site_cell}{warp_cells}"
            f"</tr>"
        )

    st.markdown(
        '<div style="overflow-x:auto;max-height:60vh;overflow-y:auto;">'
        f'<table class="site-url-table"><thead><tr>{hdr}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table></div>',
        unsafe_allow_html=True,
    )


# ── Excel エクスポート ────────────────────────────────────
def _build_excel(rows: list[tuple], idx: dict, months: list[int]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URLマトリクス"

    fills = {
        "filled":     PatternFill("solid", fgColor="1A2A1A"),
        "filled_new": PatternFill("solid", fgColor="1A253A"),
        "pending":    PatternFill("solid", fgColor="3A3210"),
        "404":        PatternFill("solid", fgColor="3A1818"),
        "dash":       PatternFill("solid", fgColor="2A1A1A"),
    }
    hdr_fill = PatternFill("solid", fgColor="1E1E2E")
    hdr_font = Font(bold=True, color="CDD6F4")
    link_font = Font(color="7C83FD", underline="single")
    text_font = Font(color="CDD6F4")

    headers = ["No.", "機関", "カテゴリ", "入札方式"] + [f"{m}月" for m in months]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    prev = None
    for row_no, (aid, aname, cat, bid, *_) in enumerate(rows, 1):
        ri = row_no + 1
        ws.cell(ri, 1, row_no).alignment = Alignment(horizontal="right")
        ws.cell(ri, 2, aname if aid != prev else "")
        ws.cell(ri, 3, cat)
        ws.cell(ri, 4, bid)
        prev = aid
        for col_off, mm in enumerate(months, 5):
            rec = idx.get((aid, aname, cat, bid, mm))
            if rec is None:
                c = ws.cell(ri, col_off, "－")
                c.fill = fills["dash"]
            else:
                state = rec["state"]
                if state == "dash":
                    c = ws.cell(ri, col_off, "－")
                else:
                    fname = (rec["filename"] or "?")[:35]
                    c = ws.cell(ri, col_off, fname)
                    if rec["url"]:
                        c.hyperlink = rec["url"]
                        c.font = link_font
                    else:
                        c.font = text_font
                c.fill = fills.get(state, fills["dash"])
            c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    for col_idx in range(5, 5 + len(months)):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.divider()
st.download_button(
    label="📥 Excel ダウンロード",
    data=_build_excel(row_keys, cell_index, FY_MONTHS),
    file_name="url_matrix_status.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
