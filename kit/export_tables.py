"""procurement.db を CSV / XLSX に書き出すスタンドアロン出力ツール。

依存は **標準ライブラリ + openpyxl のみ**（pandas/streamlit不要）。
DB再構築後（kit/rebuild_all.py 完了後）に実行して、表計算で使える形に吐き出す。

出力:
  CSV   : out/contracts.csv（全契約。要求元・7本柱コードをJOIN済み）
          + out/by_agency/<category>.csv（機関カテゴリ別）
  XLSX  : out/contracts.xlsx
          - "サマリー" シート（機関×FY件数・金額、7本柱別）
          - "契約一覧" シート（最大10万行。超過時はCSVへ誘導）
          列幅・ヘッダー固定・金額カンマ書式つき

実行:
  python kit/export_tables.py                      # 既定: data/db/procurement.db → kit/out/
  python kit/export_tables.py --db path.db --out dir
  python kit/export_tables.py --format csv         # csv だけ / xlsx だけ
  python kit/export_tables.py --fy 2024 2025       # 年度で絞る
  python kit/export_tables.py --agency-category 海上自衛隊
"""
from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB = PROJECT_ROOT / "data" / "db" / "procurement.db"
DEFAULT_OUT = PROJECT_ROOT / "kit" / "out"

# 出力列（contracts + enrichment JOIN）。順序がそのまま列順になる
COLUMNS = [
    ("id", "契約ID"),
    ("agency_category", "機関カテゴリ"),
    ("agency_id", "機関ID"),
    ("agency_name", "機関名"),
    ("requesting_org", "要求元(推定)"),
    ("fiscal_year", "年度"),
    ("contract_date", "契約日"),
    ("contract_name", "契約名"),
    ("vendor_name", "業者名"),
    ("contract_amount", "契約額(円)"),
    ("estimated_price", "予定価格(円)"),
    ("award_rate", "落札率"),
    ("bid_method", "入札方式"),
    ("contract_type", "契約種別"),
    ("pillar_l1_code", "7本柱L1"),
    ("pillar_l2_code", "7本柱L2"),
    ("source_type", "出所種別"),
    ("source_url", "出所URL"),
]

BASE_SQL = """
SELECT c.id, c.agency_category, c.agency_id, c.agency_name,
       o.requesting_org,
       c.fiscal_year, c.contract_date, c.contract_name, c.vendor_name,
       c.contract_amount, c.estimated_price, c.award_rate,
       c.bid_method, c.contract_type,
       p.pillar_l1_code, p.pillar_l2_code,
       c.source_type, c.source_url
FROM contracts c
LEFT JOIN contract_pillar p ON p.contract_id = c.id
LEFT JOIN (
    SELECT contract_id, requesting_org,
           ROW_NUMBER() OVER (PARTITION BY contract_id
                              ORDER BY confidence DESC) AS rn
    FROM contract_requesting_org
) o ON o.contract_id = c.id AND o.rn = 1
"""


def connect_ro(db: Path) -> sqlite3.Connection:
    con = sqlite3.connect(f"file:{db.as_posix()}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    return con


def build_where(fys, cats) -> tuple[str, list]:
    conds, params = [], []
    if fys:
        conds.append("c.fiscal_year IN (%s)" % ",".join("?" * len(fys)))
        params += list(fys)
    if cats:
        conds.append("c.agency_category IN (%s)" % ",".join("?" * len(cats)))
        params += list(cats)
    return (" WHERE " + " AND ".join(conds)) if conds else "", params


def fetch_rows(con, fys, cats) -> list[sqlite3.Row]:
    where, params = build_where(fys, cats)
    return con.execute(BASE_SQL + where + " ORDER BY c.fiscal_year, c.agency_id, c.id",
                       params).fetchall()


# ── CSV 出力 ─────────────────────────────────────────────────────────────────

def write_csv(rows, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    keys = [k for k, _ in COLUMNS]
    headers = [h for _, h in COLUMNS]

    main = out_dir / "contracts.csv"
    with main.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r[k] for k in keys])
    print(f"CSV: {main}  ({len(rows):,}行)")

    # 機関カテゴリ別
    by_dir = out_dir / "by_agency"
    by_dir.mkdir(exist_ok=True)
    groups: dict[str, list] = {}
    for r in rows:
        groups.setdefault(r["agency_category"] or "その他", []).append(r)
    for cat, grp in sorted(groups.items()):
        safe = "".join(ch for ch in cat if ch not in r'\/:*?"<>|')
        path = by_dir / f"{safe}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in grp:
                w.writerow([r[k] for k in keys])
    print(f"CSV(機関別): {by_dir}/ ({len(groups)}カテゴリ)")


# ── XLSX 出力（openpyxl）─────────────────────────────────────────────────────

def write_xlsx(con, rows, out_dir: Path, fys, cats) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARN: openpyxl 未導入のため XLSX をスキップ（pip install openpyxl）")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    where, params = build_where(fys, cats)

    # サマリーシート
    ws = wb.active
    ws.title = "サマリー"
    ws.append(["機関×年度 集計"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["機関カテゴリ", "年度", "件数", "契約額合計(億円)"])
    for c in ws[2]:
        c.fill, c.font = hdr_fill, hdr_font
    for r in con.execute(
            "SELECT agency_category, fiscal_year, COUNT(*), "
            "ROUND(SUM(contract_amount)/1e8,1) FROM contracts c" + where +
            " GROUP BY agency_category, fiscal_year "
            "ORDER BY agency_category, fiscal_year", params):
        ws.append(list(r))

    ws.append([])
    start = ws.max_row + 1
    ws.append(["7本柱(L1)別 集計"])
    ws[f"A{start}"].font = Font(bold=True, size=14)
    ws.append(["7本柱L1", "件数", "契約額合計(億円)"])
    for c in ws[start + 1]:
        c.fill, c.font = hdr_fill, hdr_font
    for r in con.execute(
            "SELECT p.pillar_l1_code, COUNT(*), ROUND(SUM(c.contract_amount)/1e8,1) "
            "FROM contracts c LEFT JOIN contract_pillar p ON p.contract_id=c.id"
            + where + " GROUP BY p.pillar_l1_code ORDER BY p.pillar_l1_code", params):
        ws.append([r[0] if r[0] is not None else "未分類", r[1], r[2]])

    # 契約一覧シート（10万行上限）
    LIMIT = 100_000
    ws2 = wb.create_sheet("契約一覧")
    ws2.append([h for _, h in COLUMNS])
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
    ws2.freeze_panes = "A2"
    keys = [k for k, _ in COLUMNS]
    amt_cols = {i + 1 for i, (k, _) in enumerate(COLUMNS)
                if k in ("contract_amount", "estimated_price")}
    for r in rows[:LIMIT]:
        ws2.append([r[k] for k in keys])
    for ci in amt_cols:
        col = get_column_letter(ci)
        for cell in ws2[col][1:]:
            cell.number_format = "#,##0"
    if len(rows) > LIMIT:
        ws2.append([f"※{len(rows):,}行中 先頭{LIMIT:,}行のみ。全件は contracts.csv 参照"])

    # 列幅
    widths = {"契約名": 50, "業者名": 30, "機関名": 22, "出所URL": 40,
              "要求元(推定)": 14, "契約日": 12, "契約額(円)": 16}
    for ci, (_, h) in enumerate(COLUMNS, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = widths.get(h, 12)

    path = out_dir / "contracts.xlsx"
    wb.save(path)
    print(f"XLSX: {path}  (契約一覧{min(len(rows), LIMIT):,}行 + サマリー)")


def main() -> None:
    ap = argparse.ArgumentParser(description="DB → CSV / XLSX 出力（標準lib+openpyxl）")
    ap.add_argument("--db", default=str(DEFAULT_DB))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--format", choices=["both", "csv", "xlsx"], default="both")
    ap.add_argument("--fy", nargs="*", type=int, help="年度で絞る")
    ap.add_argument("--agency-category", nargs="*", help="機関カテゴリで絞る")
    args = ap.parse_args()

    db = Path(args.db)
    if not db.exists():
        sys.exit(f"DBがありません: {db}（先に kit/rebuild_all.py を実行）")
    out_dir = Path(args.out)

    con = connect_ro(db)
    rows = fetch_rows(con, args.fy, args.agency_category)
    if not rows:
        sys.exit("該当データなし（絞り込み条件を確認）")
    total_oku = sum((r["contract_amount"] or 0) for r in rows) / 1e8
    print(f"対象: {len(rows):,}件 / {total_oku:,.0f}億円")

    if args.format in ("both", "csv"):
        write_csv(rows, out_dir)
    if args.format in ("both", "xlsx"):
        write_xlsx(con, rows, out_dir, args.fy, args.agency_category)
    con.close()
    print(f"SUMMARY rows={len(rows)} out={out_dir}")


if __name__ == "__main__":
    main()
