"""Teams/SharePoint 配布用 Excel ダッシュボードを生成する（標準lib + openpyxl）。

「Excel for the web でブラウザ閲覧」を前提に、二層構成の .xlsx を1ファイル出力:
  - "ダッシュボード"  集計グラフ（機関別/年度推移/7本柱/上位ベンダー）
  - "集計"           ピボット相当の集計表（AutoFilter付き）
  - "明細"           全契約（AutoFilter＋ヘッダー固定。ブラウザでも絞り込める）

PivotTable/スライサーは openpyxl では安定生成できないため、
ネイティブ棒・折れ線グラフ + AutoFilter で「壊れない・Web表示で動く」ダッシュボードにする。
（真のピボット＋スライサーが要る場合は手順書のテンプレ方式を参照）

依存: 標準ライブラリ + openpyxl のみ（pandas不要）。

実行:
  python kit/make_dashboard_xlsx.py                       # 既定DB → kit/out/防衛調達ダッシュボード.xlsx
  python kit/make_dashboard_xlsx.py --db <repaired.db> --out <dir>
  python kit/make_dashboard_xlsx.py --no-detail           # 明細シートを省く（さらに軽量）
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB = PROJECT_ROOT / "data" / "db" / "procurement.db"
DEFAULT_OUT = PROJECT_ROOT / "kit" / "out"

PILLAR_NAME = {
    1: "P1 スタンド・オフ", 2: "P2 統合防空", 3: "P3 無人アセット",
    4: "P4 領域横断", 5: "P5 指揮統制", 6: "P6 機動展開",
    7: "P7 持続性・強靱性", 8: "P8 防衛生産基盤",
}

DETAIL_COLS = [
    ("agency_category", "機関カテゴリ"), ("agency_id", "機関ID"),
    ("agency_name", "機関名"), ("fiscal_year", "年度"),
    ("contract_date", "契約日"), ("contract_name", "契約名"),
    ("vendor_name", "業者名"), ("contract_amount", "契約額(円)"),
    ("award_rate", "落札率"), ("bid_method", "入札方式"),
    ("pillar_l1_code", "7本柱"), ("source_url", "出所URL"),
]


def connect_ro(db: Path) -> sqlite3.Connection:
    con = sqlite3.connect(f"file:{db.as_posix()}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    return con


def main() -> None:
    ap = argparse.ArgumentParser(description="Teams配布用Excelダッシュボード生成")
    ap.add_argument("--db", default=str(DEFAULT_DB))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--no-detail", action="store_true", help="明細シートを省く")
    args = ap.parse_args()

    db = Path(args.db)
    if not db.exists():
        sys.exit(f"DBがありません: {db}")

    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl が必要です: pip install openpyxl")

    con = connect_ro(db)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    HDR = PatternFill("solid", fgColor="1F3864")
    HF = Font(color="FFFFFF", bold=True)
    TITLE = Font(bold=True, size=16, color="1F3864")
    YEN = "#,##0"

    wb = Workbook()

    # ── 集計シート（グラフのデータ源）──────────────────────────────────────
    agg = wb.active
    agg.title = "集計"

    def block(ws, start_row, title, header, rows):
        ws.cell(start_row, 1, title).font = Font(bold=True, size=13)
        hr = start_row + 1
        for ci, h in enumerate(header, 1):
            c = ws.cell(hr, ci, h); c.fill = HDR; c.font = HF
        r = hr + 1
        for row in rows:
            for ci, v in enumerate(row, 1):
                ws.cell(r, ci, v)
            r += 1
        return hr, r - 1  # header_row, last_data_row

    # 機関カテゴリ別（金額・件数）
    cat_rows = [(r[0] or "その他", r[1], round((r[2] or 0) / 1e8, 1))
                for r in con.execute(
        "SELECT agency_category, COUNT(*), SUM(contract_amount) "
        "FROM contracts GROUP BY agency_category ORDER BY SUM(contract_amount) DESC")]
    cat_h, cat_last = block(agg, 1, "機関カテゴリ別", ["機関", "件数", "金額(億円)"], cat_rows)

    # 年度推移
    fy_start = cat_last + 3
    fy_rows = [(r[0], r[1], round((r[2] or 0) / 1e8, 1)) for r in con.execute(
        "SELECT fiscal_year, COUNT(*), SUM(contract_amount) "
        "FROM contracts GROUP BY fiscal_year ORDER BY fiscal_year")]
    fy_h, fy_last = block(agg, fy_start, "年度推移", ["年度", "件数", "金額(億円)"], fy_rows)

    # 7本柱別
    pl_start = fy_last + 3
    pl_rows = [(PILLAR_NAME.get(r[0], "未分類"), r[1], round((r[2] or 0) / 1e8, 1))
               for r in con.execute(
        "SELECT p.pillar_l1_code, COUNT(*), SUM(c.contract_amount) "
        "FROM contracts c LEFT JOIN contract_pillar p ON p.contract_id=c.id "
        "GROUP BY p.pillar_l1_code ORDER BY p.pillar_l1_code")]
    pl_h, pl_last = block(agg, pl_start, "7本柱別（独自分類）", ["7本柱", "件数", "金額(億円)"], pl_rows)

    # 上位ベンダー15（㈱/（株）を株式会社へ寄せて名寄せ）
    vn_start = pl_last + 3
    vn_rows = [(r[0], r[1], round((r[2] or 0) / 1e8, 1)) for r in con.execute(
        "SELECT REPLACE(REPLACE(REPLACE(vendor_name,'㈱','株式会社'),'（株）','株式会社'),'(株)','株式会社') AS v, "
        "COUNT(*), SUM(contract_amount) FROM contracts "
        "WHERE vendor_name IS NOT NULL AND vendor_name NOT LIKE '%未記載%' "
        "GROUP BY v ORDER BY SUM(contract_amount) DESC LIMIT 15")]
    vn_h, vn_last = block(agg, vn_start, "上位ベンダー15", ["業者", "件数", "金額(億円)"], vn_rows)

    for col in "ABC":
        agg.column_dimensions[col].width = 26 if col == "A" else 14
    agg.column_dimensions["A"].width = 30

    # ── ダッシュボードシート（グラフ）──────────────────────────────────────
    dash = wb.create_sheet("ダッシュボード", 0)
    dash.sheet_view.showGridLines = False
    dash["A1"] = "防衛調達ダッシュボード"
    dash["A1"].font = Font(bold=True, size=22, color="1F3864")
    dash["A2"] = (f"出典: 防衛省公表データ（財計第2017号）を独自集計 / "
                  f"生成: {datetime.now():%Y-%m-%d} / 全{con.execute('SELECT COUNT(*) FROM contracts').fetchone()[0]:,}件")
    dash["A2"].font = Font(size=11, color="808080")

    def bar(title, hrow, lastrow, anchor, color="4472C4"):
        ch = BarChart(); ch.type = "bar"; ch.title = title; ch.height = 8; ch.width = 16
        data = Reference(agg, min_col=3, min_row=hrow, max_row=lastrow)
        cats = Reference(agg, min_col=1, min_row=hrow + 1, max_row=lastrow)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ch.legend = None
        dash.add_chart(ch, anchor)

    def line(title, hrow, lastrow, anchor):
        ch = LineChart(); ch.title = title; ch.height = 8; ch.width = 16
        data = Reference(agg, min_col=3, min_row=hrow, max_row=lastrow)
        cats = Reference(agg, min_col=1, min_row=hrow + 1, max_row=lastrow)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ch.legend = None
        dash.add_chart(ch, anchor)

    bar("機関カテゴリ別 契約額（億円）", cat_h, cat_last, "A4")
    line("年度推移 契約額（億円）", fy_h, fy_last, "J4")
    bar("7本柱別 契約額（億円）", pl_h, pl_last, "A22")
    bar("上位ベンダー15 契約額（億円）", vn_h, vn_last, "J22")

    # ── 明細シート ─────────────────────────────────────────────────────────
    if not args.no_detail:
        det = wb.create_sheet("明細")
        det.append([h for _, h in DETAIL_COLS])
        for c in det[1]:
            c.fill = HDR; c.font = HF
        det.freeze_panes = "A2"
        keys = [k for k, _ in DETAIL_COLS]
        sql = """
        SELECT c.agency_category, c.agency_id, c.agency_name, c.fiscal_year,
               c.contract_date, c.contract_name, c.vendor_name, c.contract_amount,
               c.award_rate, c.bid_method, p.pillar_l1_code, c.source_url
        FROM contracts c LEFT JOIN contract_pillar p ON p.contract_id=c.id
        ORDER BY c.fiscal_year, c.agency_id, c.id
        """
        n = 0
        amt_idx = keys.index("contract_amount") + 1
        for r in con.execute(sql):
            det.append([r[k] for k in keys]); n += 1
        col = get_column_letter(amt_idx)
        for cell in det[col][1:]:
            cell.number_format = YEN
        det.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{n+1}"
        widths = {"契約名": 50, "業者名": 30, "機関名": 22, "出所URL": 38, "契約日": 12}
        for ci, (_, h) in enumerate(DETAIL_COLS, 1):
            det.column_dimensions[get_column_letter(ci)].width = widths.get(h, 12)
        print(f"明細: {n:,}行")

    path = out_dir / "防衛調達ダッシュボード.xlsx"
    wb.save(path)
    con.close()
    mb = path.stat().st_size / 1e6
    print(f"SUMMARY out={path} size={mb:.1f}MB sheets={wb.sheetnames}")
    if mb > 100:
        print("⚠️ 100MB超: Excel for the web で開けません。--no-detail で軽量化を")


if __name__ == "__main__":
    main()
